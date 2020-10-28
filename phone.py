import openpyxl
import sys
import os

def main ():
    wb = get_excel_doc()
    sheet = get_sheet(wb)

def get_excel_doc():
    file_obj = {}
    os.chdir('./')
    files = os.listdir()
    file_map = {}
    count = 1
    for file in files:
        if file[-5:] == '.xlsx':
            file_map[count] = file
            count += 1
    doc_range = range(1,len(file_map)+1)
    print(f"Excel Documents: {file_map}")
    get_doc = input("Pick document number: ")
    while int(get_doc) not in doc_range:
        print("Not a valid document")
        get_doc = input("Pick document number: ")
    file_obj["wb"] = openpyxl.load_workbook(f"{file_map[int(get_doc)]}")
    file_obj["name"] = f"{file_map[int(get_doc)]}"
    return file_obj

def get_sheet(wb):
    sheet_range = range(1, len(wb["wb"].sheetnames)+1)
    print("Sheets: ", end="")
    for i in range(len(wb["wb"].sheetnames)):
        print(f"{i+1}) {wb['wb'].sheetnames[i]}   ", end="")
    print("")
    pick_sheet = input("Pick sheet number: ")
    while int(pick_sheet) not in sheet_range:
        print("Not a valid sheet")
        pick_sheet = input("Pick sheet number: ")
    return wb["wb"][wb["wb"].sheetnames[int(pick_sheet)-1]]


if __name__ == "__main__":
    main()